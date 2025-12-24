import argparse
import json
import os
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
from openpyxl.styles import Alignment, Font


def _safe_float(v: Any) -> Optional[float]:
    try:
        if v is None:
            return None
        return float(v)
    except Exception:
        return None


def _unwrap_cmd_result(resp: Any) -> Dict[str, Any]:
    """
    Accepts a JSON-RPC response envelope produced by our test harness and returns:
      { ok, result, ... }
    """
    cur = resp
    for _ in range(6):
        if isinstance(cur, dict) and "ok" in cur and "result" in cur:
            return cur
        if isinstance(cur, dict) and "result" in cur:
            cur = cur.get("result")
            continue
        break
    return {}


def _ensure_sheet(wb: openpyxl.Workbook, name: str) -> openpyxl.worksheet.worksheet.Worksheet:
    if name in wb.sheetnames:
        ws_old = wb[name]
        wb.remove(ws_old)
    ws = wb.create_sheet(title=name)
    return ws


def _write_kv(ws: openpyxl.worksheet.worksheet.Worksheet, row: int, key: str, value: Any) -> int:
    ws.cell(row=row, column=1, value=key).font = Font(bold=True)
    ws.cell(row=row, column=2, value=value)
    return row + 1


def _format_wrap(ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
    wrap = Alignment(wrap_text=True, vertical="top")
    for row in ws.iter_rows():
        for c in row:
            if c.value is None:
                continue
            c.alignment = wrap


def _build_segment_rows(ctx: Dict[str, Any]) -> Tuple[List[List[Any]], float]:
    loops = ctx.get("loops") or []
    wall_ids_by_seg = ctx.get("wallIdsBySegmentKey") or {}
    ceilings_by_seg = ctx.get("ceilingIdsBySegmentKey") or {}

    walls = ctx.get("walls") or []
    wall_id_to_type: Dict[int, str] = {}
    for w in walls:
        try:
            wid = int(w.get("wallId") or 0)
            if wid <= 0:
                continue
            wall_id_to_type[wid] = str(w.get("typeName") or "")
        except Exception:
            continue

    ceiling_id_to_height: Dict[int, float] = {}
    ceilings = ctx.get("ceilings") or []
    for c in ceilings:
        try:
            cid = int(c.get("elementId") or 0)
            h = _safe_float(c.get("heightFromRoomLevelMm"))
            if cid > 0 and h is not None:
                ceiling_id_to_height[cid] = float(h)
        except Exception:
            continue

    rows: List[List[Any]] = []
    total_len = 0.0

    for loop in loops:
        segs = (loop or {}).get("segments") or []
        for seg in segs:
            li = int(seg.get("loopIndex") or 0)
            si = int(seg.get("segmentIndex") or 0)
            key = f"{li}:{si}"

            st = seg.get("start") or {}
            en = seg.get("end") or {}

            length = _safe_float(seg.get("lengthMm")) or 0.0
            total_len += float(length)

            boundary_kind = seg.get("boundaryKind") or ""

            # Segment ceiling heights (prefer per-segment baked values, else map ids->heights)
            seg_ceiling_heights = seg.get("ceilingHeightsFromRoomLevelMm")
            heights: List[float] = []
            if isinstance(seg_ceiling_heights, list) and seg_ceiling_heights:
                for hv in seg_ceiling_heights:
                    h = _safe_float(hv)
                    if h is not None:
                        heights.append(float(h))
            else:
                ids = ceilings_by_seg.get(key) or []
                if isinstance(ids, list):
                    for cid in ids:
                        try:
                            cid_int = int(cid)
                            if cid_int in ceiling_id_to_height:
                                heights.append(ceiling_id_to_height[cid_int])
                        except Exception:
                            continue

            heights = sorted(set(round(h, 3) for h in heights))
            heights_text = "\n".join(str(h) for h in heights) if heights else ""

            wall_ids = wall_ids_by_seg.get(key) or []
            wall_ids_int: List[int] = []
            if isinstance(wall_ids, list):
                for wid in wall_ids:
                    try:
                        wall_ids_int.append(int(wid))
                    except Exception:
                        continue
            wall_type_names = [wall_id_to_type.get(wid, "") for wid in wall_ids_int if wid in wall_id_to_type]
            wall_type_names = [x for x in wall_type_names if x]
            wall_type_names_sorted = sorted(set(wall_type_names))

            rows.append(
                [
                    si,
                    li,
                    st.get("x"),
                    st.get("y"),
                    st.get("z"),
                    en.get("x"),
                    en.get("y"),
                    en.get("z"),
                    round(float(length), 3),
                    heights_text,
                    boundary_kind,
                    len(wall_ids_int),
                    len(wall_type_names_sorted),
                    "\n".join(wall_type_names_sorted),
                ]
            )

    return rows, total_len


def main() -> int:
    ap = argparse.ArgumentParser(description="Export standardized SegmentWallTypes sheet from compare JSON.")
    ap.add_argument("json_path", help="Path to test_room_finish_takeoff_context_compare_*.json")
    ap.add_argument(
        "--xlsx",
        dest="xlsx_path",
        default="",
        help="Output xlsx path (default: same basename as json_path). If exists, only SegmentWallTypes is replaced.",
    )
    args = ap.parse_args()

    json_path = os.path.abspath(args.json_path)
    if not os.path.exists(json_path):
        raise SystemExit(f"JSON not found: {json_path}")

    xlsx_path = args.xlsx_path
    if not xlsx_path:
        base, _ = os.path.splitext(json_path)
        xlsx_path = base + ".xlsx"
    xlsx_path = os.path.abspath(xlsx_path)

    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    resp = (data.get("responses") or {}).get("get_room_finish_takeoff_context") or {}
    cmd = _unwrap_cmd_result(resp)
    if not cmd.get("ok"):
        raise SystemExit("get_room_finish_takeoff_context is not ok in this JSON.")
    ctx = cmd.get("result") or {}

    if os.path.exists(xlsx_path):
        wb = openpyxl.load_workbook(xlsx_path)
    else:
        wb = openpyxl.Workbook()
        # Remove the default sheet to avoid clutter.
        try:
            wb.remove(wb.active)
        except Exception:
            pass

    ws = _ensure_sheet(wb, "SegmentWallTypes")

    # Title
    ws.cell(row=1, column=1, value="Room boundary segments → matched wall type names").font = Font(
        bold=True, size=12
    )

    room = ctx.get("room") or {}
    metrics = ctx.get("metrics") or {}

    r = 3
    r = _write_kv(ws, r, "roomId", room.get("roomId"))
    r = _write_kv(ws, r, "roomName", room.get("name"))
    r = _write_kv(ws, r, "levelId", room.get("levelId"))
    r = _write_kv(ws, r, "levelName", room.get("levelName"))

    # Base offset (room)
    r = _write_kv(ws, r, "baseOffsetMm", room.get("baseOffsetMm"))
    r = _write_kv(ws, r, "baseOffsetDisplay", room.get("baseOffsetDisplay"))

    # Perimeter: param + measured
    r = _write_kv(ws, r, "perimeterParamDisplay", metrics.get("perimeterParamDisplay"))
    r = _write_kv(ws, r, "perimeterParamMm", metrics.get("perimeterParamMm"))
    r = _write_kv(ws, r, "perimeterMeasuredMm", metrics.get("perimeterMm"))

    # Room height (global)
    r = _write_kv(ws, r, "roomHeightMm", metrics.get("roomHeightMm"))

    # Floors/Ceilings summary
    floors = ctx.get("floors") or []
    ceilings = ctx.get("ceilings") or []

    def fmt_floor(x: Dict[str, Any]) -> str:
        return (
            f'{x.get("elementId")} (typeId={x.get("typeId")}) : {x.get("typeName")} '
            f'@ top {x.get("topHeightFromRoomLevelMm")}mm'
        )

    def fmt_ceiling(x: Dict[str, Any]) -> str:
        return (
            f'{x.get("elementId")} (typeId={x.get("typeId")}) : {x.get("typeName")} '
            f'@ {x.get("heightFromRoomLevelMm")}mm'
        )

    floors_text = "\n".join(fmt_floor(x) for x in floors) if floors else ""
    ceilings_text = "\n".join(fmt_ceiling(x) for x in ceilings) if ceilings else ""

    r = _write_kv(ws, r, "floors (id:type @ topHeightFromRoomLevelMm)", floors_text)
    r = _write_kv(ws, r, "ceilings (id:type @ heightFromRoomLevelMm)", ceilings_text)

    # Segment table
    rows, total_len = _build_segment_rows(ctx)
    r += 1

    # Length checks
    perim_measured = _safe_float(metrics.get("perimeterMm"))
    perim_param = _safe_float(metrics.get("perimeterParamMm"))

    r = _write_kv(ws, r, "segmentsLengthTotalMm", round(total_len, 3))
    if perim_measured is not None:
        r = _write_kv(ws, r, "segmentsMinusPerimeterMeasuredMm", round(total_len - float(perim_measured), 3))
    if perim_param is not None:
        r = _write_kv(ws, r, "segmentsMinusPerimeterParamMm", round(total_len - float(perim_param), 3))

    tol = 5.0
    note = "OK"
    if perim_measured is not None and abs(total_len - float(perim_measured)) > tol:
        note = (
            f"WARN: segments total differs from metrics.perimeterMm by >{tol}mm. "
            "Rounding or boundary options may be involved."
        )
    r = _write_kv(ws, r, "lengthCheckNote", note)

    r += 1
    header_row = r
    headers = [
        "segmentIndex",
        "loopIndex",
        "startXmm",
        "startYmm",
        "startZmm",
        "endXmm",
        "endYmm",
        "endZmm",
        "lengthMm",
        "roomHeightMm (ceiling heights from room level)",
        "boundaryKind",
        "matchedWallCount",
        "matchedWallTypeCount",
        "matchedWallTypeNames",
    ]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=r, column=c, value=h)
        cell.font = Font(bold=True)
    r += 1

    for row in rows:
        for c, v in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=v)
        r += 1

    # Simple column widths
    widths = {
        1: 12,
        2: 10,
        3: 12,
        4: 12,
        5: 12,
        6: 12,
        7: 12,
        8: 12,
        9: 10,
        10: 28,
        11: 12,
        12: 16,
        13: 18,
        14: 40,
    }
    for col, w in widths.items():
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

    _format_wrap(ws)

    try:
        wb.save(xlsx_path)
        print(f"Updated: {xlsx_path} (SegmentWallTypes)")
    except PermissionError:
        base, ext = os.path.splitext(xlsx_path)
        alt = base + "_updated" + ext
        n = 2
        while os.path.exists(alt):
            alt = f"{base}_updated_{n}{ext}"
            n += 1
        wb.save(alt)
        print(f"Updated (saved as new file because the target is locked): {alt} (SegmentWallTypes)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
